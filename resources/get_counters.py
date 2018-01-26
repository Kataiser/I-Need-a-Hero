import openpyxl
import csv

from resources import namenum_converter as conv
from resources import exception_handler


def get_counter_xlsx(hero1, hero2):
    x = conv.strip_dead(hero1)
    y = conv.strip_dead(hero2)
    if x == y:
        return 0

    posx = hero_reference.index(x) + 2  # because python lists start at 0
    posy = hero_reference.index(y) + 2  # and the counter table starts at 2

    result = sheet.cell(row=posx, column=posy).value  # load from spreadsheet
    if not result:  # probably useless
        result = 0
    return int(-result)


def get_counter_csv(hero1, hero2):
    hero1 = conv.fancify(hero1)
    hero2 = conv.fancify(hero2)
    if hero1 == 'Soldier 76':
        hero1 = 'Soldier: 76'
    if hero2 == 'Soldier 76':
        hero2 = 'Soldier: 76'
    if hero1 == hero2:
        return 0

    counter_value1 = None
    counter_value2 = None
    for full_counter in counters_table:
        counter_first = full_counter[0]
        counter_second = full_counter[1]
        if counter_first == hero1 and counter_second == hero2:
            counter_value1 = full_counter[2]
        if counter_first == hero2 and counter_second == hero1:
            counter_value2 = full_counter[2]
        if counter_value1 and counter_value2:
            break
    if counter_value1 and counter_value2:
        counter_value1 = int(convert_values_table[str(counter_value1)])
        counter_value2 = int(convert_values_table_reversed[str(counter_value2)])
        counter_value_average = (counter_value1 + counter_value2) / 2
        return -counter_value_average
    return None  # if doomfist or orisa


def get_counter_rb(hero1, hero2):
    if hero1 == 'unknown' or hero1 == 'loading':
        return 0
    if hero2 == 'unknown' or hero2 == 'loading':
        return 0
    if hero1 == hero2:
        return 0

    try:
        counter_value = rb_counters_dict[hero1][hero2]
        return counter_value[0]
    except KeyError:
        return 0


def get_counter(hero1, hero2):
    if hero1 == 'unknown' or hero1 == 'loading':
        return 0
    if hero2 == 'unknown' or hero2 == 'loading':
        return 0

    xlsx_results = get_counter_xlsx(hero1, hero2)
    csv_results = get_counter_csv(hero1, hero2)
    rb_results = get_counter_rb(hero1, hero2)  # weird still
    if csv_results is None:
        return xlsx_results + rb_results
    else:
        return (xlsx_results + csv_results) / 2 + rb_results


def get_synergy(hero1, hero2, blank_is_negative):
    hero1 = conv.fancify(hero1)
    hero2 = conv.fancify(hero2)
    if hero1 == 'Soldier 76':  # this is formatted with a colon in the csv
        hero1 = 'Soldier: 76'  # but without one in this program
    if hero2 == 'Soldier 76':
        hero2 = 'Soldier: 76'
    if blank_is_negative:
        if 'unknown' in hero1 or 'loading' in hero1:
            return -2
        if 'unknown' in hero2 or 'loading' in hero2:
            return -2
    if hero1 == hero2:
        return 0

    synergy_value1 = None
    synergy_value2 = None
    for full_synergy in synergies_table:
        synergy_first = full_synergy[0]
        synergy_second = full_synergy[1]
        if synergy_first == hero1 and synergy_second == hero2:
            synergy_value1 = full_synergy[2]
        if synergy_first == hero2 and synergy_second == hero1:
            synergy_value2 = full_synergy[2]
        if synergy_value1 and synergy_value2:
            break
    if synergy_value1 and synergy_value2:
        synergy_value1 = int(convert_values_table[str(synergy_value1)])
        synergy_value2 = int(convert_values_table[str(synergy_value2)])
        synergy_value_average = (synergy_value1 + synergy_value2) / 2
        return -synergy_value_average
    return 0  # if doomfist or orisa


exception_handler.setup_excepthook()

# setup for xlsx
wb = openpyxl.load_workbook('resources/counters.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

hero_reference = []

for i in range(2, 27):
    to_add = sheet.cell(row=i, column=1).value.lower()  # builds list of heroes
    hero_reference.append(to_add)

hero_reference.append('unknown')  # because these aren't in the spreadsheet
hero_reference.append('loading')

# setup for csv
with open('resources/counters.csv', newline='') as csvfile:  # loads the counters database
    reader = csv.DictReader(csvfile)
    counters_table = []
    for line in reader:  # builds a better counter table, as a list of tuples
        counter_tuple = (line['agent'], line['target'], line['median'])
        counters_table.append(counter_tuple)

with open('resources/rb_counters.csv', newline='') as csvfile:
    reader = csv.reader(csvfile)
    rb_counters_dict = {}
    for line in reader:
        if line[0] not in rb_counters_dict:
            rb_counters_dict[line[0]] = {}
        rb_counters_dict[line[0]][line[1]] = [int(line[2])]

convert_values_table = {'5': -2, '4': -1, '3': 0, '2': 1, '1': 2, '0': 0}
convert_values_table_reversed = {'5': 2, '4': 1, '3': 0, '2': -1, '1': -2, '0': 0}

with open('resources/synergies.csv', newline='') as csvfile:  # loads the synergies database
    reader = csv.DictReader(csvfile)
    synergies_table = []
    for line in reader:  # builds a synergy table
        synergy_tuple = (line['agent'], line['target'], line['median'])
        synergies_table.append(synergy_tuple)

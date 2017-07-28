import openpyxl
import namenum_converter as conv


def get_counter(hero1, hero2):
    if hero1 == 'unknown' or hero1 == 'loading':
        return 0
    if hero2 == 'unknown' or hero2 == 'loading':
        return 0

    x = conv.strip_dead(hero1)
    y = conv.strip_dead(hero2)
    if x == y:
        return 0

    posx = hero_reference.index(x) + 2  # because python lists start at 0
    posy = hero_reference.index(y) + 2  # and the counter table starts at 2

    result = sheet.cell(row=posx, column=posy).value  # load from spreadsheet
    if not result:  # probably useless
        result = 0
    return int(result)

wb = openpyxl.load_workbook('counters.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

hero_reference = []

for i in range(2, 27):
    to_add = sheet.cell(row=i, column=1).value.lower()  # builds list of heroes
    hero_reference.append(to_add)

hero_reference.append('unknown')  # because these aren't in the spreadsheet
hero_reference.append('loading')
